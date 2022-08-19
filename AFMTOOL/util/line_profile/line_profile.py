import openpyxl

from openpyxl.utils import get_column_letter
import numpy as np
import matplotlib.pyplot as plt

#Assumes contacts are in horizontal array
def plot_line_profile(filename_formatted, array, vert_line, cu_sh_width, x, y, r, scan_size, scan_pixels_len):
    
    avg_height = np.mean(array)
    
    #Adjusting for coordinate differences between picture and numpy array
    #Convert from image scale to array index
    x_index = int(x*scan_pixels_len/768)
    y_index = int(y*scan_pixels_len/768)
    r_index = int(r*scan_pixels_len/768)
    
    #Take average within band of width r/2
    if not vert_line:
        #Taking horizontal band
        # Uncomment to show band used
        # array[max(0,int(y_index-r_index/2)): min(scan_pixels_len,int(y_index+r_index/2)), :]=0
        # plt.imshow(array, cmap="copper")
        # plt.show()
        line_data = array[max(0,int(y_index-r_index/2)): min(scan_pixels_len,int(y_index+r_index/2)), :]
        line_data = np.mean(line_data, axis=0)
    else:
        #Vertical band
        line_data = array[:, max(0,int(x_index-r_index/2)): min(scan_pixels_len,int(x_index+r_index/2))]
        line_data = np.mean(line_data, axis=1).flatten()
        
    
    
    fig_x = np.linspace(0,scan_size,scan_pixels_len)
    fig = plt.figure(figsize=(18,12))
    ax = plt.subplot(111)
    ax.plot(fig_x, line_data, color ="blue")
    ax.set_ylabel('nm', fontsize=10)
    ax.set_xlabel('μm', fontsize=10)
    
    #Plot vertical lines denoting copper contact
    r_um = r*scan_size/768 #Get radius in um
    if not vert_line:
        x_um = x*scan_size/768
    else:
        #TODO: Change naming
        x_um = y*scan_size/768
    #Take range slightly smaller than r to be safe
    left_lim = x_um-r_um*cu_sh_width
    right_lim = x_um+r_um*cu_sh_width
    ax.plot((left_lim, left_lim), (min(line_data), 0.1+max(line_data)), color='black', linestyle='-.')
    ax.plot((right_lim, right_lim), (min(line_data), 0.1+max(line_data)), color='black', linestyle='-.')
    #Uncomment to see avg_height
    #ax.plot((0,19), (avg_height, avg_height), color='black', linestyle='-.')
    #avg_line = np.mean(line_data)
    #ax.plot((0,19), (avg_line, avg_line), color='red', linestyle='-.')
    
    left_lim_index = int(left_lim*scan_pixels_len/scan_size)
    right_lim_index = int(right_lim*scan_pixels_len/scan_size)
    avg_copper_height = np.mean(line_data[int(left_lim_index):int(right_lim_index)])
    
    dishing = True if (avg_copper_height<avg_height) else False
    
    #Plot vertical lines denoting polymer
    #Note that polymer limits are shifted inwards (towards each other) by 0.2*r from the 
    #calculated "ideal" positions, to be safe not to touch regions of copper contacts in case 
    #centers or radius of circles identified are not accurate 
    if(x_um>=10): 
        pol_right_lim = x_um-r_um*1.2
        ax.plot((pol_right_lim, pol_right_lim), (min(line_data), 0.1+max(line_data)), color='red', linestyle='-.')
        #Note that find_pol_limit gives index, need to convert to um
        
        #Used for calculation of roll off
        cut_off_limit = find_pol_limit(line_data, int((pol_right_lim)*scan_pixels_len/scan_size), avg_height, avg_copper_height, dishing, False, int(r_um*scan_pixels_len/scan_size))
        cut_off_height = line_data[cut_off_limit]
        ax.plot(((scan_size/scan_pixels_len)*cut_off_limit, (scan_size/scan_pixels_len)*cut_off_limit), (min(line_data), 0.1+max(line_data)), color='green', linestyle='--')
        pol_left_lim = (scan_size/scan_pixels_len)*cut_off_limit+0.2*r_um
        ax.plot((pol_left_lim, pol_left_lim), (min(line_data), 0.1+max(line_data)), color='red', linestyle='-.')
    else:
        pol_left_lim = x_um+r_um*1.2
        ax.plot((pol_left_lim, pol_left_lim), (min(line_data), 0.1+max(line_data)), color='red', linestyle='-.')
        cut_off_limit = find_pol_limit(line_data, int((pol_left_lim)*scan_pixels_len/scan_size), avg_height, avg_copper_height, dishing, True, int(r_um*scan_pixels_len/scan_size))
        cut_off_height = line_data[cut_off_limit]
        ax.plot(((scan_size/scan_pixels_len)*cut_off_limit, (scan_size/scan_pixels_len)*cut_off_limit), (min(line_data), 0.1+max(line_data)), color='green', linestyle='--')
        pol_right_lim = (scan_size/scan_pixels_len)*cut_off_limit-0.2*r_um
        ax.plot((pol_right_lim, pol_right_lim), (min(line_data), 0.1+max(line_data)), color='red', linestyle='-.')
    img_path = "../AFMTOOL/line_profile_imgs/"+str(filename_formatted)+"line_plot"
    
    #Plot roll off lines
    pol_center = int((scan_pixels_len/scan_size)*(pol_left_lim+pol_right_lim)/2)
    pol_center_height = line_data[pol_center]
    ax.plot(((scan_size/scan_pixels_len)*pol_center, (scan_size/scan_pixels_len)*pol_center), (min(line_data), 0.1+max(line_data)), color='green', linestyle='--')
    plt.savefig(img_path, bbox_inches='tight')
    
    plt.close(fig)
    avg_pol_height = np.mean(line_data[int(pol_left_lim*scan_pixels_len/scan_size):int(pol_right_lim*scan_pixels_len/scan_size)])
    #Calculate roll off
    
    roll_off = pol_center_height - cut_off_height
    step_height = avg_copper_height-avg_pol_height
    return step_height, pol_left_lim, pol_right_lim, roll_off
    
def insert_line_profile(filename_formatted, excel_file_path, col_num, step_height, roll_off):
    
    step_height = "Error: could not obtain step height" if step_height is None else step_height
    roll_off = "Error: could not obtain roll_off" if roll_off is None else roll_off
    
    img_path  = "../AFMTOOL/line_profile_imgs/"+str(filename_formatted)+"line_plot.png"
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb["Sheet"]
    
    col_letter = get_column_letter(col_num+1)
    
    line_img = openpyxl.drawing.image.Image(img_path)
    line_img.height = 93
    line_img.width = 140
    
    line_img.anchor = col_letter + '10'
    ws.add_image(line_img)
    
    #Insert step height
    ws[col_letter + '9'] = "{:.3f}".format(step_height)
    ws[col_letter + '11'] = "{:.3f}".format(roll_off)
    
    wb.save(excel_file_path)
    

def find_pol_limit(array, first_limit, avg_height, avg_copper_height, dishing, search_right, r_pixel):

    index = first_limit

    if(dishing):
        if(search_right):
            #+15 so starting finding closer to the center of polymer, to avoid roll off regions
            #Move left/right until closer to average copper height than average height
            #Index will stop at the border in the circle
            index +=r_pixel
            while(index<255 and not all(abs(i-avg_copper_height)<abs(i-avg_height) for i in array[index:index+5])):
                index+=1
        else:
            index-=r_pixel
            while(index>=0 and not all(abs(i-avg_copper_height)<abs(i-avg_height) for i in array[index-5:index])):
                index-=1
    else:
        if(search_right):
            index+=r_pixel
            while(index<255 and not all(abs(i-avg_copper_height)<abs(i-avg_height) for i in array[index:index+5])):
                index+=1
        else:
            index-=r_pixel
            while(index>=0 and not all(abs(i-avg_copper_height)<abs(i-avg_height) for i in array[index-5:index])):
                index-=1
    
    return index



    