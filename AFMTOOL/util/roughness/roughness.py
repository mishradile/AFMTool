""" 
Function to calculate roughness(Ra) of a square region
Inputs- (x,y) coordinate the center of square and L, side length of square, in micrometer. 

Machine-learning algorithm in util/shape_recogniser will give the function the center of the Cu connections for input. 
Side length of square will be set by user

From https://www.olympus-ims.com/en/metrology/surface-roughness-measurement-portal/parameters/#!cms[focus]=007, 
Roughness Average (Ra) is given by average of absolute values of deviation from the mean of a given sample. 
In Numpy's terminology it's Mean Absolute Deviation from mean (https://www.geeksforgeeks.org/absolute-deviation-and-absolute-mean-deviation-using-numpy-python/)
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

from numpy import mean, absolute
from matplotlib.patches import Rectangle
import matplotlib.pyplot as plt
from datetime import datetime
import pytz

tz_SG = pytz.timezone('Asia/Singapore') 
datetime_SG = datetime.now(tz_SG)
#For saving files with timestamps 
format_timestring = datetime_SG.strftime("%m%d%Y%H%M")

def find_ra(array, detected_circles):
    """ 
    Returns average roughness of region of copper contact centered at (x,y), 
    with default side length of region = 13 pixels = 1.016 micrometer for copper,
    and 25 pixels for polymer, 
    so that the region extends 6 pixels/12 pixels in each direction from the center piece. 
    
    Coordinates in numpy array generated by pySPM matches orientation of image generated
    
    For calculation of polymer roughness, assumed that diamerter:pitch ratio at least 1:2. We'll find 
    roughness of polymer vector (r,r) away from each center of circle
    """
    #Counter for sum of ra across all detected contacts
    total_ra=0
    pol_total_ra =0
    circles_count=0
    pol_area_count =0
    take_bottom_left = False
    for pt in detected_circles[0, :]:
        circles_count+=1
        x,y, r = int(pt[0]*256/768), int(pt[1]*256/768), int(pt[2]*256/768)
    
        
        #Define sample area to calculate roughness 
        #used min, max in case selected center is too close to the borders
        #TODO: Record in doc coordinate system 
        sample = array[max(0, y-6): min(256, y+6+1), max(0,x-6):min(256, x+6+1)]
        #Uncomment to show area used
        # array[max(0, y-6): min(256, y+6+1), max(0,x-6):min(256, x+6+1)] =0
        # plt.imshow(array, cmap="copper")
        # plt.show()
        total_ra+=mean(absolute(sample - mean(sample)))
        
        #Polymer roughness 
        x_pol = x+2*r
        y_pol = y-2*r
        #Check at least a portion of sample area is in range
        if(y_pol+12>0 and (x_pol)-12<256):
            pol_sample = array[max((y_pol)-12,0):min((y_pol)+12+1, 256), max(x_pol-12,0):min(x_pol+12+1, 256)] 
            pol_total_ra +=mean(absolute(pol_sample - mean(pol_sample)))
            pol_area_count+=1
        
    if(circles_count==0):
        copper_ra = -1
    else:
        copper_ra = total_ra/circles_count

    if(pol_area_count==0):
        #If all polymer area out of bound, try take bottom left of each circle
        for pt in detected_circles[0, :]:
            x,y, r = int(pt[0]*256/768), int(pt[1]*256/768), int(pt[2]*256/768)
            #Polymer roughness 
            x_pol = x-2*r
            y_pol = y+2*r
            if(y_pol-12<256 and (x_pol)-12<256):
                pol_sample = array[max((y_pol)-12,0):min((y_pol)+12+1, 256), max(x_pol-12,0):min(x_pol+12+1, 256)] 
                pol_total_ra +=mean(absolute(pol_sample - mean(pol_sample)))
                pol_area_count+=1
        if (pol_area_count==0):
            pol_ra = "Errored: Please calculate polymer roughness manually."
        else:
            pol_ra = pol_total_ra/pol_area_count
            take_bottom_left = True
    else:
        pol_ra = pol_total_ra/pol_area_count
        
    print([copper_ra, pol_ra])
        
    return [copper_ra, pol_ra, take_bottom_left]

def insert_ra(excel_file_path, ra, pol_ra, col_num):
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb["Sheet"]
    
    col_letter = get_column_letter(col_num+1)
    try:
        ws[col_letter+'7'] = "{:.3f}".format(ra)
        ws[col_letter+'8'] = "{:.3f}".format(pol_ra)
    except ValueError:
        ws[col_letter+'7'] = ra
        ws[col_letter+'8'] = pol_ra
    wb.save(excel_file_path)
    
    return 0
    


def insert_ref_image(filename_formatted, excel_file_path, col_num):
    
    img_path  =  "../results/ref_regions_imgs/"+str(filename_formatted)+"ref_plot.png"
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb["Sheet"]
    
    col_letter = get_column_letter(col_num+1)
    
    line_img = openpyxl.drawing.image.Image(img_path)
    line_img.height = 140
    line_img.width = 140
    
    line_img.anchor = col_letter + '11'
    ws.add_image(line_img)
    
    wb.save(excel_file_path)