from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Alignment, Side
import openpyxl
import pytz
from openpyxl.utils import get_column_letter

workbook = Workbook()
sheet = workbook.active

tz_SG = pytz.timezone('Asia/Singapore') 
datetime_SG = datetime.now(tz_SG)
#For saving files with timestamps 
format_timestring = datetime_SG.strftime("%m%d%Y%H%M")

blueFill = PatternFill(start_color='000066CC',
                   end_color='000066CC',
                   fill_type='solid')

def create_xl_template():
    """ 
    Creates empty template Excel sheet to be filled in 
    Return path to created Excel sheet
    """
    
    sheet["A1"] = "Lot ID: "
    sheet["A3"] = "File Name:"
    sheet["A5"] = "Top View"
    sheet["A6"] = "3D View"
    sheet["A7"] = "Cu Roughness(nm) 1umX1um"
    sheet["A8"] = "Po roughness (nm) 2umX2um"
    sheet["A9"] = "Step height (nm) Protusion : + Dishing:-"

    sheet["A10"] = "Line Profile Image"
    sheet["A11"] = "Roll-off"
    sheet["A12"] = "Reference Image"
    
    sheet["B1"] = "Slot ID: "
    
    sheet["C1"] = "CMP recipe: "
    
    

    sheet["A5"].fill = blueFill
    sheet["A6"].fill = blueFill
    sheet["A7"].fill = blueFill
    sheet["A8"].fill = blueFill
    sheet["A9"].fill = blueFill
    sheet["A10"].fill = blueFill
    sheet["A11"].fill = blueFill
    sheet["A12"].fill = blueFill
    
    
    
    #Set row height for images
    sheet.row_dimensions[5].height = 104
    sheet.row_dimensions[6].height = 104
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[10].height = 70
    sheet.row_dimensions[12].height = 104
    
    sheet.column_dimensions['G'].width = 20

            
    sheet.column_dimensions['A'].width = 25

    target_path = "../results/xlSheets/"+ format_timestring+ "_AFMResults.xlsx"
    workbook.save(target_path)
    
    return target_path

def insert_xl(excel_file_path, img_path_2d, img_path_3d, col_num):
    
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb["Sheet"]
    
    
    img_path_2d = img_path_2d[11:]+".png"
    img_path_3d = img_path_3d[11:]+".png"
    #images/3_2d_plot.png
    img_2d = openpyxl.drawing.image.Image(img_path_2d)
    img_3d = openpyxl.drawing.image.Image(img_path_3d)
    img_2d.height = 140
    img_2d.width = 140
    img_3d.height = 140
    img_3d.width = 140
    
    col_letter = get_column_letter(col_num+1)
    ws.column_dimensions[col_letter].width = 20
    img_2d.anchor =  col_letter+'5'
    img_3d.anchor =  col_letter+'6'
    
    ws[col_letter+'3'] = img_path_2d[7:-12]
    
    ws[col_letter+'3'].fill = blueFill
    ws.add_image(img_2d)
    ws.add_image(img_3d)
    
    wb.save(excel_file_path)

def style_excel_final(excel_file_path):
    
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb["Sheet"]
    
     #Align texts and style cell borders
    double = Side(border_style="medium", color="000000")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment= Alignment(wrap_text=True, vertical='center', horizontal = 'center')
            cell.border = Border(top=double, left=double, right=double, bottom=double)
    
    
    wb.save(excel_file_path)

    